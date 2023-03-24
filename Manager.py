import configparser
import threading
import time
import tkinter
from os import _exit, getcwd,listdir
from tkinter.filedialog import askdirectory, askopenfilename
from tkinter.messagebox import askyesno, showinfo
from typing import Callable, Union
import customtkinter
import openpyxl
import pystray
from DatabaseModel import ICON_PATH, DBConnection, Sources, Table, pd
from PIL import Image
from pystray import MenuItem as item
from tktimepicker import SpinTimePickerModern, constants

customtkinter.set_appearance_mode("Light")
customtkinter.set_default_color_theme("blue")

class IntSpinbox(customtkinter.CTkFrame):
    def __init__(self, *args,
                 width: int = 100,
                 height: int = 32,
                 step_size: int = 1,
                 command: Callable = None,
                 **kwargs):
        super().__init__(*args, width=width, height=height, **kwargs)

        self.step_size = step_size
        self.command = command

        self.configure(fg_color=("gray78", "gray28"))  # set frame color

        self.grid_columnconfigure((0, 2), weight=0)  # buttons don't expand
        self.grid_columnconfigure(1, weight=1)  # entry expands

        self.subtract_button = customtkinter.CTkButton(self, text="-", width=height-6, height=height-6,
                                                       command=self.subtract_button_callback)
        self.subtract_button.grid(row=0, column=0, padx=(3, 0), pady=3)

        self.entry = customtkinter.CTkEntry(self, width=width-(2*height), height=height-6, border_width=0)
        self.entry.grid(row=0, column=1, columnspan=1, padx=3, pady=3, sticky="ew")

        self.add_button = customtkinter.CTkButton(self, text="+", width=height-6, height=height-6,
                                                  command=self.add_button_callback)
        self.add_button.grid(row=0, column=2, padx=(0, 3), pady=3)

        # default value
        self.entry.insert(0, "0.0")

    def add_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            value = int(self.entry.get()) + self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
        except ValueError:
            return

    def subtract_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            value = int(self.entry.get()) - self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
        except ValueError:
            return

    def get(self) -> Union[int, None]:
        try:
            return int(self.entry.get())
        except ValueError:
            return None

    def set(self, value):
        self.entry.delete(0, "end")
        self.entry.insert(0, str(value))

class UpdateMenu(customtkinter.CTkToplevel):
    def file_selection(self):
        
        filetypes = (
            ('Excel', ('*.csv','*.xls', '*.xlsx')),
        )

        self.file = askopenfilename(
            title='Open files',
            initialdir='/',
            filetypes=filetypes)

        self.filename = self.file.split("/")[-1]

        self.fentry.insert(index=0,string=self.filename.split("/")[-1])

        workbook = openpyxl.load_workbook(filename=self.file)
        self.sheetList = workbook.sheetnames
        self.sheet_menu.configure(require_redraw=True,values=self.sheetList)
    
    def __init__(self, dbc : DBConnection, iid,**kwargs):
        super().__init__(**kwargs)
        self.dbc = dbc
        self.iid = iid
        self.dbc._session()

        self.CurrentRecord = self.dbc.sl.query(Sources).filter(Sources.RowId == iid)

        self.title("Update Sources")
        self.geometry(f"{1035}x{400}")

        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)


        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)

        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame,text="Manage Sources",font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.savebtn = customtkinter.CTkButton(self.sidebar_frame, command=self._save,text="Save")
        self.savebtn.grid(row=1, column=0, padx=20, pady=10)
        self.backbtn = customtkinter.CTkButton(self.sidebar_frame, command=self.destroy,text="Back")
        self.backbtn.grid(row=2, column=0, padx=20, pady=10)


        self.tabview = customtkinter.CTkTabview(self, width=25)
        self.tabview.grid(row=0,rowspan=3, column=1, padx=(20, 20), pady=(20, 20), sticky="nsew")

        if self.CurrentRecord.first().Method == "FTP":
            self.UpdateFTP()
        elif self.CurrentRecord.first().Method == "API":
            self.UpdateAPI()

    def UpdateAPI(self):
        self.tabview.add("Add API Source")
        self.API_frame = customtkinter.CTkFrame(self.tabview.tab("Add API Source"))
        self.API_frame.grid(row=1, column=1, columnspan=3,padx=(20, 20), pady=(20, 20), sticky="ew")

        
        self.TableNameEntry = customtkinter.CTkEntry(master=self.API_frame, placeholder_text="Table Name",width=200)
        self.TableNameEntry.grid(row=2,column=1, padx=(20, 20), pady=(20, 20), sticky="nsew")
        self.TableNameEntry.insert(index=0,string=self.CurrentRecord.first().Table_Name)

        self.EndpointEntry = customtkinter.CTkEntry(master=self.API_frame, placeholder_text="Endpoint Url",width=200)
        self.EndpointEntry.grid(row=2,column=2, padx=(20, 20), pady=(20, 20), sticky="nsew")
        self.EndpointEntry.insert(index=0,string=self.CurrentRecord.first().Endpoint_Link)

        self.AuthMenu = customtkinter.CTkOptionMenu(self.API_frame, dynamic_resizing=False, values=['Microsoft','Crowdstrike','Forcepoint','Duo Security','VmWare'],width=200)
        self.AuthMenu.grid(row=2, column=3, padx=20, pady=(20, 20))
        self.AuthMenu.set(self.CurrentRecord.first().Auth)
        
        self.FrequncyMenu = customtkinter.CTkOptionMenu(self.API_frame, dynamic_resizing=False, values=['Daily','Weekly','Monthly'],width=200)
        self.FrequncyMenu.grid(row=3, column=1, padx=20, pady=(20, 20))
        self.FrequncyMenu.set(self.CurrentRecord.first().Frequency)

        self.DayMenu = IntSpinbox(self.API_frame, width=200, step_size=1)
        self.DayMenu.grid(row=3, column=2, padx=20, pady=(20, 20))
        self.DayMenu.set(self.CurrentRecord.first().Day)


        self.TimeFrame = customtkinter.CTkFrame(master=self.API_frame,width=200,bg_color='white')
        self.TimeFrame.grid(row=3, column=3, padx=20, pady=(20, 20))
        self.TimePicker = SpinTimePickerModern(parent=self.TimeFrame,orient=constants.HORIZONTAL,per_orient=constants.HORIZONTAL)
        self.TimePicker.addAll(constants.HOURS24)
        self.TimePicker.pack(expand=True, fill="both")

        self.CreateAPI = customtkinter.CTkButton(master=self.API_frame, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"),text="Create Record" ,command=self.createAPIRecord,width=200)
        self.CreateAPI.grid(row=4, column=1, padx=20, pady=(20, 20))
        ...
    
    def UpdateFTP(self):
        self.tabview.add("Add FTP Source")
        self.FTP_frame = customtkinter.CTkFrame(self.tabview.tab("Add FTP Source"))
        self.FTP_frame.grid(row=1, column=1, columnspan=3,padx=(20, 20), pady=(20, 20), sticky="ew")


        self.FTP_File = customtkinter.CTkButton(master=self.FTP_frame, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"),text="Select Credentails File" ,command=self.file_selection,width=200)
        self.FTP_File.grid(row=3, column=1, padx=20, pady=(20, 20))

        self.fentry = customtkinter.CTkEntry(master=self.FTP_frame,width=200)
        self.fentry.grid(row=3, column=2,padx=(20, 20), pady=(20, 20), sticky="new")
        self.fentry.insert(index=0,string=self.CurrentRecord.first().File_Name)

        self.sheet_menu = customtkinter.CTkOptionMenu(self.FTP_frame, dynamic_resizing=False, values=self.sheetList,width=200)
        self.sheet_menu.grid(row=4, column=1, padx=20, pady=(20, 20))
        self.sheet_menu.set(value=self.CurrentRecord.first().Sheet)

        self.SkipRowsEntry = customtkinter.CTkEntry(master=self.FTP_frame, placeholder_text="Skip Rows",width=200)
        self.SkipRowsEntry.grid(row=4,column=2, padx=(20, 20), pady=(20, 20), sticky="nsew")
        self.SkipRowsEntry.insert(index=0,string=self.CurrentRecord.first().Skip_Rows)

        self.CreateFTP = customtkinter.CTkButton(master=self.FTP_frame, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"),text="Create Record" ,command=self.createFTPRecord,width=200)
        self.CreateFTP.grid(row=5, column=1, padx=20, pady=(20, 20))
        ...

    def _save(self):
        self.dbc.sl.commit()
        showinfo(title='Source Updated',message='Selected Source was Updated.')
        self.destroy()
        ...

    def createFTPRecord(self):
        sr = 0 if self.SkipRowsEntry.get() == "" else int(self.SkipRowsEntry.get())
        newRecord = {'Table_Name':self.fentry.get(),'File_Name':self.filename,'Method':'FTP','Sheet':self.sheet_menu.get(),'Skip_Rows' : sr}
        self.CurrentRecord.update(newRecord)
        ...

    def createAPIRecord(self):
        d = 1 if self.DayMenu.get() > 28 else self.DayMenu.get()
        newRecord = {'Table_Name':self.TableNameEntry.get(),'Endpoint_Link':self.EndpointEntry.get(),'Method':'API','Frequency':self.FrequncyMenu.get(),'Day' : d,'Time':str(self.TimePicker.hours()) + ":" + str(self.TimePicker.minutes()),'Auth':self.AuthMenu.get()}
        self.CurrentRecord.update(newRecord)

    ...

class SourceMenu(customtkinter.CTkToplevel):
    def file_selection(self):
        
        filetypes = (
            ('Excel', ('*.csv','*.xls', '*.xlsx')),
        )

        self.file = askopenfilename(
            title='Open files',
            initialdir='/',
            filetypes=filetypes)

        self.filename = self.file.split("/")[-1]

        self.fentry.insert(index=0,string=self.filename.split("/")[-1])

        workbook = openpyxl.load_workbook(filename=self.file)
        self.sheetList = workbook.sheetnames
        self.sheet_menu.configure(require_redraw=True,values=self.sheetList)

    def __init__(self, dbc : DBConnection):
        self.dbc = dbc
        self.sheetList = ['DefaultSheet']
        self.dbc._session()

        super().__init__()

        self.title("Manage Sources")
        self.geometry(f"{1035}x{400}")

        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)


        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)

        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame,text="Manage Sources",font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.savebtn = customtkinter.CTkButton(self.sidebar_frame, command=self._save,text="Save")
        self.savebtn.grid(row=1, column=0, padx=20, pady=10)
        self.backbtn = customtkinter.CTkButton(self.sidebar_frame, command=self.destroy,text="Back")
        self.backbtn.grid(row=2, column=0, padx=20, pady=10)


        self.tabview = customtkinter.CTkTabview(self, width=25)
        self.tabview.grid(row=0,rowspan=3, column=1, padx=(20, 20), pady=(20, 20), sticky="nsew")
        self.tabview.add("Add API Source")
        self.tabview.add("Add FTP Source")


        # API
        self.API_frame = customtkinter.CTkFrame(self.tabview.tab("Add API Source"))
        self.API_frame.grid(row=1, column=1, columnspan=3,padx=(20, 20), pady=(20, 20), sticky="ew")

        
        self.TableNameEntry = customtkinter.CTkEntry(master=self.API_frame, placeholder_text="Table Name",width=200)
        self.TableNameEntry.grid(row=2,column=1, padx=(20, 20), pady=(20, 20), sticky="nsew")

        self.EndpointEntry = customtkinter.CTkEntry(master=self.API_frame, placeholder_text="Endpoint Url",width=200)
        self.EndpointEntry.grid(row=2,column=2, padx=(20, 20), pady=(20, 20), sticky="nsew")

        self.AuthMenu = customtkinter.CTkOptionMenu(self.API_frame, dynamic_resizing=False, values=['Microsoft','Crowdstrike','Forcepoint','Duo Security','VmWare'],width=200)
        self.AuthMenu.grid(row=2, column=3, padx=20, pady=(20, 20))
        
        self.FrequncyMenu = customtkinter.CTkOptionMenu(self.API_frame, dynamic_resizing=False, values=['Daily','Weekly','Monthly'],width=200)
        self.FrequncyMenu.grid(row=3, column=1, padx=20, pady=(20, 20))

        self.DayMenu = IntSpinbox(self.API_frame, width=200, step_size=1)
        self.DayMenu.grid(row=3, column=2, padx=20, pady=(20, 20))
        self.DayMenu.set(1)


        self.TimeFrame = customtkinter.CTkFrame(master=self.API_frame,width=200,bg_color='white')
        self.TimeFrame.grid(row=3, column=3,padx=10, pady=(20, 20),columnspan=2)
        self.TimePicker = SpinTimePickerModern(parent=self.TimeFrame,orient=constants.HORIZONTAL,per_orient=constants.HORIZONTAL)
        self.TimePicker.addAll(constants.HOURS24)
        self.TimePicker.pack(expand=True, fill="both")

        self.CreateAPIRecord = customtkinter.CTkButton(master=self.API_frame, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"),text="Create Record" ,command=self.createAPIRecord,width=200)
        self.CreateAPIRecord.grid(row=4, column=1, padx=20, pady=(20, 20))


        # FTP
        self.FTP_frame = customtkinter.CTkFrame(self.tabview.tab("Add FTP Source"))
        self.FTP_frame.grid(row=1, column=1, columnspan=3,padx=(20, 20), pady=(20, 20), sticky="ew")


        self.FTP_File = customtkinter.CTkButton(master=self.FTP_frame, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"),text="Select File" ,command=self.file_selection,width=200)
        self.FTP_File.grid(row=3, column=1, padx=20, pady=(20, 20))

        self.fentry = customtkinter.CTkEntry(master=self.FTP_frame, placeholder_text="Table Name / Selected File",width=200)
        self.fentry.grid(row=3, column=2,padx=(20, 20), pady=(20, 20), sticky="new")

        self.sheet_menu = customtkinter.CTkOptionMenu(self.FTP_frame, dynamic_resizing=False, values=self.sheetList,width=200)
        self.sheet_menu.grid(row=4, column=1, padx=20, pady=(20, 20))

        self.SkipRowsEntry = customtkinter.CTkEntry(master=self.FTP_frame, placeholder_text="Skip Rows",width=200)
        self.SkipRowsEntry.grid(row=4,column=2, padx=(20, 20), pady=(20, 20), sticky="nsew")

        self.CreateFTPRecord = customtkinter.CTkButton(master=self.FTP_frame, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"),text="Create Record" ,command=self.createFTPRecord,width=200)
        self.CreateFTPRecord.grid(row=5, column=1, padx=20, pady=(20, 20))

    def _save(self):
        self.dbc.sl.commit()
        showinfo(title='Source Added',message='New Source was added.')
        self.destroy()
        ...

    def createFTPRecord(self):
        sr = 0 if self.SkipRowsEntry.get() == "" else int(self.SkipRowsEntry.get())
        newRecord = {'Table_Name':self.fentry.get(),'File_Name':self.filename,'Method':'FTP','Sheet':self.sheet_menu.get(),'Skip_Rows' : sr}
        newFTPRecord = Sources(**newRecord)
        self.dbc.sl.add(newFTPRecord)
        ...

    def createAPIRecord(self):
        d = 1 if self.DayMenu.get() > 28 else self.DayMenu.get()
        newRecord = {'Table_Name':self.TableNameEntry.get(),'Endpoint_Link':self.EndpointEntry.get(),'Method':'API','Frequency':self.FrequncyMenu.get(),'Day' : d,'Time':str(self.TimePicker.hours()) + ":" + str(self.TimePicker.minutes()),'Auth':self.AuthMenu.get()}
        newAPIRecord = Sources(**newRecord)
        self.dbc.sl.add(newAPIRecord)

class SettingsMenu(customtkinter.CTkToplevel):
    config = configparser.ConfigParser()

    def directory_selection(self):
        self.dbox.configure(state='normal')
        self.newdirectory = askdirectory(initialdir = getcwd(),title = "Set Directory for Manual File updates.")
        self.dbox.insert(index='0.0',text=self.newdirectory)
        self.dbox.configure(state='disabled')
        ...

    def __init__(self,parent : customtkinter.CTk,app_settings : configparser.SectionProxy):
        self.current_app_Settings = app_settings
        self.parent = parent
        super().__init__()

        self.title("Settings")
        self.geometry(f"{835}x{400}")


        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)


        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)

        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame,text="Settings",font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.savebtn = customtkinter.CTkButton(self.sidebar_frame, command=self._save,text="Save")
        self.savebtn.grid(row=1, column=0, padx=20, pady=10)
        self.backbtn = customtkinter.CTkButton(self.sidebar_frame, command=self.destroy,text="Back")
        self.backbtn.grid(row=2, column=0, padx=20, pady=10)


        self.tabview = customtkinter.CTkTabview(self, width=25)
        self.tabview.grid(row=0,rowspan=3, column=1, padx=(20, 20), pady=(20, 20), sticky="nsew")
        self.tabview.add("App Configuration")


        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark"], command=self.parent.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=6, column=0, padx=20, pady=(10, 10))

        self.scaling_label = customtkinter.CTkLabel(self.sidebar_frame, text="UI Scaling:", anchor="w")
        self.scaling_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"], command=self.parent.change_scaling_event)
        self.scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 20))


        # Configuration Screen

        self.ServerInstance = customtkinter.CTkTextbox(self.tabview.tab("App Configuration"), width=350,height=2)
        self.ServerInstance.grid(row=1, column=2, columnspan=2, padx=(20, 0), pady=(20, 20), sticky="nsew")
        self.DatabaseInstance = customtkinter.CTkTextbox(self.tabview.tab("App Configuration"), width=350,height=2)
        self.DatabaseInstance.grid(row=2, column=2, columnspan=2, padx=(20, 0), pady=(20, 20), sticky="nsew")

        self.ServerInstance.insert("0.0",text=app_settings["Server"])
        self.ServerInstance.configure(state='disabled')
        self.DatabaseInstance.insert("0.0",text=app_settings["Database"])
        self.DatabaseInstance.configure(state='disabled')

        self.monitoringDirectory = customtkinter.CTkButton(master=self.tabview.tab("App Configuration"), fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"),text="Change directory" ,command=self.directory_selection)
        self.monitoringDirectory.grid(row=3, column=2, padx=20, pady=(20, 20))

        self.dbox = customtkinter.CTkTextbox(self.tabview.tab("App Configuration"), width=350,height=2)
        self.dbox.grid(row=3, column=3,padx=(20, 20), pady=(20, 20), sticky="nsew")
        self.dbox.insert(index="0.0",text=app_settings["Path"])
        self.dbox.configure(state="disabled")

    def _save(self):
        self.config.add_section('App')
        self.config.set('App', 'Server', self.current_app_Settings['Server'])
        self.config.set('App', 'Database', self.current_app_Settings['Database'])
        self.config.set('App', 'Path',  self.newdirectory)

        with open(r"config.cfg", 'w') as configfile:
            self.config.write(configfile)
        configfile.close()

        self.destroy()

class Startup(customtkinter.CTk):
    config = configparser.ConfigParser()
    def __init__(self):
        super().__init__()

        self.config.read(['config.cfg'])
        self.app_settings = self.config['App']
        files = listdir(str(getcwd()))
        if '.env' in files:
            self.db = DBConnection(ServerInstance=self.app_settings['Server'],Database=self.app_settings['Database'],directory=self.app_settings['Path'],connection="NotTrusted")
        else:
            self.db = DBConnection(ServerInstance=self.app_settings['Server'],Database=self.app_settings['Database'],directory=self.app_settings['Path'])
        self.menu = (item('Show', self.show_window),item('Quit', self.__exit))
        self.db._connect()

        self.APISourcesdata= pd.read_sql_query(sql="SELECT [RowId],[Table_Name],[Frequency],[Day],[Time],[Auth] FROM [Sources] WHERE [Method] = 'API';",con=self.db.connectionInstance)
        self.FTPSourcesdata= pd.read_sql_query(sql="SELECT [RowId],[Table_Name],[File_Name],[Sheet],[Skip_Rows] FROM [Sources] WHERE [Method] = 'FTP';",con=self.db.connectionInstance)
        self.Logsdata= pd.read_sql_table(table_name="Logs",con=self.db.connectionInstance)

        self.db.start_monitor()

        organization_logo = customtkinter.CTkImage(light_image=Image.open(r"Company.png"), dark_image=Image.open(r"Company.png"), size=(150, 70))

        # configure window
        self.title("Monitor Manager")
        self.iconbitmap('appicon.ico')
        
        self.geometry(f"{1535}x{765}")

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)


        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame,text="",font=customtkinter.CTkFont(size=20, weight="bold"),image=organization_logo)
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))


        self.sidebar_button_1 = customtkinter.CTkButton(self.sidebar_frame, command=self.openSettings,text="Settings")
        self.sidebar_button_1.grid(row=1, column=0, padx=20, pady=10)
        self.sidebar_button_2 = customtkinter.CTkButton(self.sidebar_frame, command=self.refreshApp,text="Refresh")
        self.sidebar_button_2.grid(row=2, column=0, padx=20, pady=10)
        self.sidebar_button_3 = customtkinter.CTkButton(self.sidebar_frame, text="Add Source", command = self._addS)
        self.sidebar_button_3.grid(row=3, column=0, padx=10, pady=10)

        self.sidebar_button_4 = customtkinter.CTkButton(self.sidebar_frame, command=self.__exit,text="Quit")
        self.sidebar_button_4.grid(row=8, column=0,  padx=20, pady=(10, 20))

        self.tabview = customtkinter.CTkTabview(self, width=25)
        self.tabview.grid(row=0,rowspan=3, column=1, padx=(20, 20), pady=(20, 20), sticky="nsew")
        self.tabview.add("Web Sources")
        self.tabview.add("File Sources")
        self.tabview.add("Logs")
        
        self.tabview.tab("Web Sources").grid_columnconfigure(0, weight=1)
        self.tabview.tab("Web Sources").grid_rowconfigure(0, weight=0)
        self.tabview.tab("Web Sources").grid_rowconfigure(1, weight=1)
        self.tabview.tab("Web Sources").grid_rowconfigure(2, weight=0)

        self.tabview.tab("Logs").grid_columnconfigure(0, weight=1)
        self.tabview.tab("Logs").grid_rowconfigure(0, weight=0)
        self.tabview.tab("Logs").grid_rowconfigure(1, weight=1)

        self.tabview.tab("File Sources").grid_columnconfigure(0, weight=1)
        self.tabview.tab("File Sources").grid_rowconfigure(0, weight=0)
        self.tabview.tab("File Sources").grid_rowconfigure(1, weight=1)
        self.tabview.tab("File Sources").grid_rowconfigure(2, weight=0)

        # API Sources Screen
        self.APISourcesTable = Table(root=self.tabview.tab("Web Sources"),dbc=self.db,data=self.APISourcesdata)

        self.btnGroup = customtkinter.CTkFrame(self.tabview.tab("Web Sources"))
        self.btnGroup.grid(row=4, column=0, sticky="nsew")

        self.call_button1 = customtkinter.CTkButton(self.btnGroup, text="Fetch", command = lambda : self.APISourcesTable.ImportSelected(self.APISourcesTable.tree.selection()),width=200)
        self.call_button1.grid(row=0, column=0, padx=10, pady=10)

        self.call_button2 = customtkinter.CTkButton(self.btnGroup, text="Fetch All", command = self.APISourcesTable.Importall,width=200)
        self.call_button2.grid(row=0, column=1, padx=10, pady=10)

        self.call_button4 = customtkinter.CTkButton(self.btnGroup, text="Update Source", command = lambda : self._updateS(self.APISourcesTable.tree.selection()),width=200)
        self.call_button4.grid(row=0, column=3, padx=10, pady=10)

        self.call_button5 = customtkinter.CTkButton(self.btnGroup, text="Remove Source", command =  lambda : self.APISourcesTable.removeSource(self.APISourcesTable.tree.selection()),width=200)
        self.call_button5.grid(row=0, column=4, padx=10, pady=10)


        # FTP Sources Screen

        self.FTPSourcesTable = Table(root=self.tabview.tab("File Sources"),dbc=self.db,data=self.FTPSourcesdata)

        self.btnGroup = customtkinter.CTkFrame(self.tabview.tab("File Sources"))
        self.btnGroup.grid(row=4, column=0, sticky="nsew")

        self.call_button4 = customtkinter.CTkButton(self.btnGroup, text="Update Source", command = lambda : self._updateS(self.FTPSourcesTable.tree.selection()),width=200)
        self.call_button4.grid(row=0, column=1, padx=10, pady=10)

        self.call_button5 = customtkinter.CTkButton(self.btnGroup, text="Remove Source", command =  lambda : self.FTPSourcesTable.removeSource(self.FTPSourcesTable.tree.selection()),width=200)
        self.call_button5.grid(row=0, column=2, padx=10, pady=10)


        # Logs Screen

        self.logsTable = Table(root=self.tabview.tab("Logs"),dbc=self.db,data=self.Logsdata)

        self.toplevel_window = None
        self.protocol('WM_DELETE_WINDOW', self.withdraw_window)

    def __exit(self):
        answer = askyesno(title='confirmation',
                    message='Are you sure that you want to quit?')
        if answer:
            self.db.o.stop()
            self.destroy()
            _exit(status=0)

    def openSettings(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = SettingsMenu(parent=self,app_settings=self.app_settings)  # create window if its None or destroyed
            self.toplevel_window.focus()
        else:
            self.toplevel_window.focus()

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)

    def refreshApp(self):
        self.db.o.stop()
        self.Logsdata = pd.read_sql_table(table_name="Logs",con=self.db.connectionInstance)
        self.APISourcesdata= pd.read_sql_query(sql="SELECT [RowId],[Table_Name],[Frequency],[Day],[Time],[Auth] FROM [Sources] WHERE [Method] = 'API';",con=self.db.connectionInstance)
        self.FTPSourcesdata= pd.read_sql_query(sql="SELECT [RowId],[Table_Name],[File_Name],[Sheet],[Skip_Rows] FROM [Sources] WHERE [Method] = 'FTP';",con=self.db.connectionInstance)
        self.logsTable.refresh(self.Logsdata)
        self.FTPSourcesTable.refresh(self.FTPSourcesdata)
        self.APISourcesTable.refresh(self.APISourcesdata)
        self.db.start_monitor()

    def _addS(self):
        if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
            self.toplevel_window = SourceMenu(dbc=self.db)  # create window if its None or destroyed
            self.toplevel_window.focus()
        else:
            self.toplevel_window.focus()
        ...

    def _updateS(self, selection : tuple):
        if len(selection) == 1:
            if self.toplevel_window is None or not self.toplevel_window.winfo_exists():
                self.toplevel_window = UpdateMenu(dbc=self.db,iid=selection[0])  # create window if its None or destroyed
                self.toplevel_window
                self.toplevel_window.focus()
            else:
                self.toplevel_window.focus()
            ...

    def withdraw_window(self):
        self.withdraw()
        self.icon = pystray.Icon("Infocepts", Image.open(ICON_PATH), "Monitor Manager", self.menu)
        self.icon.run()

    def show_window(self):
        self.icon.stop()
        self.deiconify()
        ...